import os
import datetime
import hashlib
import random
import io
import psycopg2
from psycopg2 import pool
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm

app = Flask(__name__)
app.secret_key = 'rahasia_kas_kelas_2026'
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(days=7)

# ================== KONFIGURASI DATABASE ==================
# GANTI URL_INI dengan Connection String Aiven kamu!
AIVEN_URL = "postgres://avnadmin:AVNS_7JPxrd03DK3rRDxj_TC@pg-213d1eb5-cworthy553-e567.g.aivencloud.com:18553/defaultdb?sslmode=require"

# Connection pool (min 1, max 5)
db_pool = psycopg2.pool.SimpleConnectionPool(1, 5, AIVEN_URL)

def get_db_connection():
    """Ambil koneksi dari pool"""
    try:
        return db_pool.getconn()
    except Exception as e:
        print(f"❌ Gagal koneksi ke database: {e}")
        raise

def close_db_connection(conn):
    """Kembalikan koneksi ke pool"""
    if conn:
        db_pool.putconn(conn)

# ================== INISIALISASI DATABASE ==================
def init_db():
    """Buat tabel dan insert data awal jika belum ada"""
    conn = get_db_connection()
    c = conn.cursor()
    
    # Tabel anggota
    c.execute('''
        CREATE TABLE IF NOT EXISTS anggota (
            id SERIAL PRIMARY KEY,
            nama TEXT UNIQUE NOT NULL,
            kelas TEXT
        )
    ''')
    
    # Tabel transaksi
    c.execute('''
        CREATE TABLE IF NOT EXISTS transaksi (
            id SERIAL PRIMARY KEY,
            anggota_id INTEGER REFERENCES anggota(id),
            tanggal TEXT NOT NULL,
            jenis TEXT NOT NULL,
            kategori TEXT,
            nominal INTEGER NOT NULL,
            keterangan TEXT
        )
    ''')
    
    # Tabel settings
    c.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    c.execute("INSERT INTO settings (key, value) VALUES ('iuran_bulanan', '20000') ON CONFLICT (key) DO NOTHING")
    
    # Tabel users
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL
        )
    ''')
    admin_pass = hashlib.sha256('Aisannous987'.encode()).hexdigest()
    c.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s) ON CONFLICT (username) DO NOTHING",
              ('admin', admin_pass, 'admin'))
    user_pass = hashlib.sha256('kelas123'.encode()).hexdigest()
    c.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s) ON CONFLICT (username) DO NOTHING",
              ('user', user_pass, 'user'))
    
    conn.commit()
    
    # ========== INSERT DAFTAR ANGGOTA ==========
    # Daftar 35 siswa (sudah diurutkan alfabet)
    siswa_list = [
        'Abdullah Ihza Dz',
        'Aditya Saputra',
        'Afikha Destianti',
        'Afiqah Febiola',
        'Arya Dika',
        'Ashifati Ashfa',
        'Basit Al Ghofur',
        'Davvi Taufiqurrahman',
        'Diaz Adnan Syahbani',
        'Eka Safrina Alifany',
        'Evalina Putri',
        'Fachri Rizqiawan',
        'Fadlan Aryo Pratama',
        'Hawri \'Ien Fadilah',
        'Inaya Savia Zahra',
        'Ismail Haniya Rahman',
        'Kayla Natasya Sofian',
        'Kirana Avrilia Sari',
        'M. Keanu Arief Sugiarto',
        'Muhammad Dimas Saputra',
        'Muhammad Faeyza Ahnaf',
        'Muhammad Zidane',
        'Nadine Greaceana Nifili',
        'Nur Aura Saparani',
        'Nur Sabrina Oktaviani',
        'Queensha Hadinata',
        'Rasyid Farhan S',
        'Reffina Rayna Puteri',
        'Ribbyna Rasya D',
        'Satrio Budi Utomo',
        'Shaula Nafeeza',
        'Syahid Arsyil',
        'Syauqiyah AJP',
        'Tanisha Firyali Hasti',
        'Tri Oktavio Sandy'
    ]
    
    # Cek apakah tabel anggota kosong
    c.execute("SELECT COUNT(*) FROM anggota")
    count = c.fetchone()[0]
    if count == 0:
        for nama in siswa_list:
            c.execute("INSERT INTO anggota (nama, kelas) VALUES (%s, %s)", (nama, 'RPL 1'))
        print(f"✅ {len(siswa_list)} anggota berhasil ditambahkan!")
    else:
        print(f"ℹ️ Sudah ada {count} anggota, tidak perlu insert ulang.")
    
    conn.commit()
    conn.close()
    print("✅ Database siap!")

# Jalankan inisialisasi saat startup
init_db()

# ================== SURPRISE ==================
SURPRISE_QUOTES = [
    "✨ Mantap! Transaksi berhasil!",
    "🎉 Yeay! Uang kas bertambah!",
    "💰 Saldo makin tebal!",
    "🌟 Kamu luar biasa!",
    "🚀 Transaksi sukses!",
    "🎊 Hore! Iuran masuk!",
    "💪 Semangat! Kas kelas sehat!"
]

def get_surprise():
    return random.choice(SURPRISE_QUOTES)

# ================== HALAMAN ==================
@app.route('/')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') == 'admin':
        return render_template('admin.html', username=session.get('username'))
    else:
        return render_template('user.html', username=session.get('username'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = c.fetchone()
        close_db_connection(conn)
        if user and user[2] == hashlib.sha256(password.encode()).hexdigest():
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['role'] = user[3]
            session.permanent = True
            return redirect(url_for('dashboard'))
        else:
            return render_template('login.html', error='Username atau Password salah!')
    return render_template('login.html', error=None)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ================== API ==================
@app.route('/api/surprise')
def api_surprise():
    return jsonify({'message': get_surprise()})

@app.route('/api/dashboard')
def api_dashboard():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT COALESCE(SUM(nominal), 0) as total FROM transaksi WHERE jenis='pemasukan'")
    total_in = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(nominal), 0) as total FROM transaksi WHERE jenis='pengeluaran'")
    total_out = c.fetchone()[0]
    saldo = total_in - total_out
    c.execute("SELECT COUNT(*) as count FROM anggota")
    total_anggota = c.fetchone()[0]
    close_db_connection(conn)
    return jsonify({
        'saldo': saldo,
        'total_in': total_in,
        'total_out': total_out,
        'total_anggota': total_anggota
    })

@app.route('/api/chart_data')
def api_chart_data():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    conn = get_db_connection()
    c = conn.cursor()
    today = datetime.date.today()
    months = []
    labels = []
    for i in range(6):
        m = today.replace(day=1) + datetime.timedelta(days=30*i)
        months.append(m.strftime('%Y-%m'))
        labels.append(m.strftime('%b %Y'))
    pemasukan = []
    pengeluaran = []
    for bulan in months:
        c.execute("SELECT COALESCE(SUM(nominal), 0) FROM transaksi WHERE jenis='pemasukan' AND tanggal LIKE %s", (bulan + '%',))
        pemasukan.append(c.fetchone()[0])
        c.execute("SELECT COALESCE(SUM(nominal), 0) FROM transaksi WHERE jenis='pengeluaran' AND tanggal LIKE %s", (bulan + '%',))
        pengeluaran.append(c.fetchone()[0])
    close_db_connection(conn)
    return jsonify({'labels': labels, 'pemasukan': pemasukan, 'pengeluaran': pengeluaran})

@app.route('/api/anggota', methods=['GET', 'POST', 'DELETE'])
def api_anggota():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    conn = get_db_connection()
    c = conn.cursor()
    if request.method == 'GET':
        c.execute("SELECT * FROM anggota ORDER BY id DESC")
        data = [{'id': row[0], 'nama': row[1], 'kelas': row[2]} for row in c.fetchall()]
        close_db_connection(conn)
        return jsonify(data)
    if session.get('role') != 'admin':
        close_db_connection(conn)
        return jsonify({'error': 'Hanya Admin'}), 403
    if request.method == 'POST':
        data = request.get_json()
        nama = data.get('nama')
        kelas = data.get('kelas', '')
        if not nama:
            close_db_connection(conn)
            return jsonify({'error': 'Nama wajib'}), 400
        try:
            c.execute("INSERT INTO anggota (nama, kelas) VALUES (%s, %s)", (nama, kelas))
            conn.commit()
            close_db_connection(conn)
            return jsonify({'message': 'Anggota ditambahkan'})
        except Exception as e:
            conn.rollback()
            close_db_connection(conn)
            return jsonify({'error': 'Nama sudah ada'}), 400
    if request.method == 'DELETE':
        data = request.get_json()
        id_anggota = data.get('id')
        if not id_anggota:
            close_db_connection(conn)
            return jsonify({'error': 'ID diperlukan'}), 400
        c.execute("DELETE FROM anggota WHERE id = %s", (id_anggota,))
        conn.commit()
        close_db_connection(conn)
        return jsonify({'message': 'Anggota dihapus'})

@app.route('/api/transaksi', methods=['GET', 'POST', 'DELETE'])
def api_transaksi():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    conn = get_db_connection()
    c = conn.cursor()
    if request.method == 'GET':
        bulan = request.args.get('bulan')
        if bulan:
            c.execute("""
                SELECT t.*, a.nama as nama_anggota 
                FROM transaksi t
                LEFT JOIN anggota a ON t.anggota_id = a.id
                WHERE t.tanggal LIKE %s
                ORDER BY t.tanggal DESC, t.id DESC
            """, (bulan + '%',))
        else:
            c.execute("""
                SELECT t.*, a.nama as nama_anggota 
                FROM transaksi t
                LEFT JOIN anggota a ON t.anggota_id = a.id
                ORDER BY t.tanggal DESC, t.id DESC
            """)
        rows = c.fetchall()
        data = []
        for row in rows:
            data.append({
                'id': row[0],
                'anggota_id': row[1],
                'tanggal': row[2],
                'jenis': row[3],
                'kategori': row[4],
                'nominal': row[5],
                'keterangan': row[6],
                'nama_anggota': row[7] if len(row) > 7 else None
            })
        close_db_connection(conn)
        return jsonify(data)
    if session.get('role') != 'admin':
        close_db_connection(conn)
        return jsonify({'error': 'Hanya Admin'}), 403
    if request.method == 'POST':
        data = request.get_json()
        anggota_id = data.get('anggota_id')
        tanggal = data.get('tanggal', datetime.date.today().isoformat())
        jenis = data.get('jenis')
        kategori = data.get('kategori', 'Lainnya')
        nominal = data.get('nominal')
        keterangan = data.get('keterangan', '')
        if not jenis or not nominal or nominal <= 0:
            close_db_connection(conn)
            return jsonify({'error': 'Data tidak valid'}), 400
        if jenis == 'pemasukan' and not anggota_id:
            close_db_connection(conn)
            return jsonify({'error': 'Pilih anggota untuk pemasukan'}), 400
        c.execute("""
            INSERT INTO transaksi (anggota_id, tanggal, jenis, kategori, nominal, keterangan)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (anggota_id, tanggal, jenis, kategori, nominal, keterangan))
        conn.commit()
        close_db_connection(conn)
        return jsonify({'message': 'Transaksi berhasil'})
    if request.method == 'DELETE':
        data = request.get_json()
        id_trans = data.get('id')
        if not id_trans:
            close_db_connection(conn)
            return jsonify({'error': 'ID diperlukan'}), 400
        c.execute("DELETE FROM transaksi WHERE id = %s", (id_trans,))
        conn.commit()
        close_db_connection(conn)
        return jsonify({'message': 'Transaksi dihapus'})

@app.route('/api/status')
def api_status():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    bulan = request.args.get('bulan', datetime.date.today().strftime('%Y-%m'))
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key='iuran_bulanan'")
    row = c.fetchone()
    target_iuran = int(row[0]) if row else 20000
    c.execute("SELECT id, nama FROM anggota")
    anggota_list = c.fetchall()
    hasil = []
    for anggota in anggota_list:
        c.execute("""
            SELECT COALESCE(SUM(nominal), 0) as total 
            FROM transaksi 
            WHERE anggota_id = %s AND jenis='pemasukan' 
            AND tanggal LIKE %s
        """, (anggota[0], bulan + '%'))
        total_bayar = c.fetchone()[0]
        status = 'Lunas' if total_bayar >= target_iuran else 'Belum'
        hasil.append({
            'id': anggota[0],
            'nama': anggota[1],
            'total_bayar': total_bayar,
            'target': target_iuran,
            'status': status
        })
    close_db_connection(conn)
    return jsonify({
        'bulan': bulan,
        'target_iuran': target_iuran,
        'data': hasil
    })

@app.route('/api/settings', methods=['GET', 'POST'])
def api_settings():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    conn = get_db_connection()
    c = conn.cursor()
    if request.method == 'GET':
        c.execute("SELECT key, value FROM settings")
        data = {row[0]: row[1] for row in c.fetchall()}
        close_db_connection(conn)
        return jsonify(data)
    if session.get('role') != 'admin':
        close_db_connection(conn)
        return jsonify({'error': 'Hanya Admin'}), 403
    if request.method == 'POST':
        data = request.get_json()
        iuran = data.get('iuran_bulanan')
        if iuran:
            c.execute("UPDATE settings SET value = %s WHERE key = 'iuran_bulanan'", (str(iuran),))
            conn.commit()
        close_db_connection(conn)
        return jsonify({'message': 'Pengaturan berhasil diubah'})

@app.route('/api/bulk_payment', methods=['POST'])
def api_bulk_payment():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json()
    tanggal = data.get('tanggal', datetime.date.today().isoformat())
    nominal = data.get('nominal')
    anggota_ids = data.get('anggota_ids', [])
    if not nominal or nominal <= 0:
        return jsonify({'error': 'Nominal tidak valid'}), 400
    if not anggota_ids:
        return jsonify({'error': 'Tidak ada anggota dipilih'}), 400
    conn = get_db_connection()
    c = conn.cursor()
    for aid in anggota_ids:
        c.execute("""
            INSERT INTO transaksi (anggota_id, tanggal, jenis, kategori, nominal, keterangan)
            VALUES (%s, %s, 'pemasukan', 'Iuran', %s, 'Bayar Cepat (Bulk)')
        """, (aid, tanggal, nominal))
    conn.commit()
    close_db_connection(conn)
    return jsonify({'message': f'✅ {len(anggota_ids)} anggota berhasil dibayar!', 'surprise': get_surprise()})

@app.route('/api/bulk_status')
def api_bulk_status():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    bulan = request.args.get('bulan', datetime.date.today().strftime('%Y-%m'))
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key='iuran_bulanan'")
    row = c.fetchone()
    target = int(row[0]) if row else 20000
    c.execute("SELECT id, nama, kelas FROM anggota ORDER BY nama")
    anggota = c.fetchall()
    hasil = []
    for a in anggota:
        c.execute("""
            SELECT COALESCE(SUM(nominal), 0) as total FROM transaksi 
            WHERE anggota_id = %s AND jenis='pemasukan' AND tanggal LIKE %s
        """, (a[0], bulan + '%'))
        total = c.fetchone()[0]
        hasil.append({
            'id': a[0],
            'nama': a[1],
            'kelas': a[2],
            'total_bayar': total,
            'target': target,
            'status': 'Lunas' if total >= target else 'Belum'
        })
    close_db_connection(conn)
    return jsonify(hasil)

# ================== EXPORT ==================
@app.route('/export/excel')
def export_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    bulan = request.args.get('bulan', datetime.date.today().strftime('%Y-%m'))
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
        SELECT t.tanggal, a.nama as anggota, t.jenis, t.kategori, t.nominal, t.keterangan
        FROM transaksi t
        LEFT JOIN anggota a ON t.anggota_id = a.id
        WHERE t.tanggal LIKE %s
        ORDER BY t.tanggal DESC
    """, (bulan + '%',))
    rows = c.fetchall()
    close_db_connection(conn)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Kas"
    headers = ['Tanggal', 'Anggota', 'Jenis', 'Kategori', 'Nominal', 'Keterangan']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row[0])
        ws.cell(row=r, column=2, value=row[1] or 'Umum')
        ws.cell(row=r, column=3, value=row[2])
        ws.cell(row=r, column=4, value=row[3] or '-')
        ws.cell(row=r, column=5, value=row[4])
        ws.cell(row=r, column=6, value=row[5] or '-')
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_size = True
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'laporan_kas_{bulan}.xlsx')

@app.route('/export/pdf')
def export_pdf():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    bulan = request.args.get('bulan', datetime.date.today().strftime('%Y-%m'))
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""
        SELECT t.tanggal, a.nama as anggota, t.jenis, t.kategori, t.nominal, t.keterangan
        FROM transaksi t
        LEFT JOIN anggota a ON t.anggota_id = a.id
        WHERE t.tanggal LIKE %s
        ORDER BY t.tanggal DESC
    """, (bulan + '%',))
    rows = c.fetchall()
    close_db_connection(conn)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=12)
    elements.append(Paragraph(f'Laporan Kas Kelas - {bulan}', title_style))
    elements.append(Spacer(1, 0.5*cm))
    data = [['Tanggal', 'Anggota', 'Jenis', 'Kategori', 'Nominal (Rp)', 'Keterangan']]
    for row in rows:
        data.append([
            row[0],
            row[1] or 'Umum',
            row[2],
            row[3] or '-',
            f"{row[4]:,}",
            row[5] or '-'
        ])
    table = Table(data, colWidths=[2*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=f'laporan_kas_{bulan}.pdf')

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)